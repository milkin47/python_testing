import openpyxl

excel_file = openpyxl.load_workbook(r"C:\Users\Milky\PycharmProjects\pythonProject\Demoexcel.xlsx")
sheet1 = excel_file.active
cellA = sheet1.cell(row=2, column=2)
print(cellA.value)

sheet1.cell(row=1, column=5).value = "Nothing"
print(sheet1.cell(row=1, column=5).value)
print("--------------")
dict = {}
for i in range(1, sheet1.max_row+1):

    for j in range(1, sheet1.max_column):
        dict[sheet1.cell(row=1, column=j).value] = sheet1.cell(row=i, column=j).value
        #print(sheet1.cell(row=i, column=j).value)
        print(dict)
    print("\n")

